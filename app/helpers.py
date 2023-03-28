import os
import openpyxl
#import pandas as pd
from pymongo import MongoClient
from functools import wraps

client = MongoClient(os.getenv("MONGO_URI"))
db = client.schema_data
service_col = db.services
medicine_col = db.medicine
diagnosis_col = db.diagnosis

urgent2k_token = os.environ.get("URGENT_2K_KEY")
#base_url = os.getenv("SAFEPAY_URL")

# decorator function frequesting api key as header
def urgent2k_token_required(f):
    @wraps(f)
    # the new, post-decoration function. Note *args and **kwargs here.
    def decorated(*args, **kwargs):
        token = None

        if 'x-access-token' in request.headers:
            token = request.headers['x-access-token']
        
        if not token:
            return {"status": False, "message": "Access token is missing at " + request.url, "data": None}, 401

        if token == urgent2k_token:
            return f(*args, **kwargs)
        else:
            return {"status": False, "message": "Invalid access token at " + request.url, "data": None}, 401

    return decorated


os.getcwd()
os.chdir("C:/Users/DELL/Downloads") #this changes our CWD, if the excel sheet is not in CWD

# file = 'PLASCHEMA Service Tariff.xlsx'
# data = pd.ExcelFile(file)
# print(data.sheet_names) #this returns the all the sheets in the excel file
# ['Sheet1']
ps = openpyxl.load_workbook('PLASCHEMA Medicine List.xlsx')
sheet = ps['Sheet1']
sheet.max_row 

# for row in range(5, sheet.max_row -9):
#     drug = sheet['B' + str(row)].value
#     dosage_form = sheet['C' + str(row)].value
#     strength = sheet['D' + str(row)].value
#     price = sheet['E' + str(row)].value

#     data = {"drug": drug, "dosage_form": dosage_form, "strength": strength, "price": price}
#     #print(data)
#     medicine_col.insert_one(data)
# print("done")

# for row in range(2, sheet.max_row + 1):
#     diagnosis = sheet['A' + str(row)].value


#     data = {"diagnosis": diagnosis}
#     diagnosis_col.insert_one(data)
# #print(data)
# print("done")

def return_services():
    data = service_col.find()
    service_list = list()
    for i in data:
        service_list.append(i.get("service"))

    print(service_list)

def return_diagnosis():
    data = diagnosis_col.find()
    diagnosis_list = list()
    for i in data:
        diagnosis_list.append(i.get("diagnosis"))

    print(diagnosis_list)
        
def return_drugs():
    data = medicine_col.find({},{ "_id": 0,"price": 0 })
    medicine_list = list()
    for i in data:
        medicine_list.append(i)

    print(medicine_list)
        

return_drugs()
